import pandas as pd
import os
from typing import Dict, List, Tuple

class GeneradorHorariosLaboratorio:
    def __init__(self):
        """
        Inicializa el Generador de Horarios de Laboratorio con configuraciones por defecto.
        """
        # Mapeo de nombres de laboratorios - EDITA ESTE DICCIONARIO para que coincida con tus nombres de laboratorio
        # Clave: Nombre del laboratorio en el archivo reporte_ocupacion
        # Valor: Nombre del laboratorio para los encabezados de salida
        self.mapeo_laboratorios = {
            'LABORATORIO GEIO CAP(25)': 'GEIO (321) TECHNE',
            'SALA DE SOFTWARE DE TECNOLOGIA E INGENIERIA DE PRODUCCION A CAP(17)': 'Sala de Software A - 16 EST - 416- TECHNE',
            'SALA DE SOFTWARE DE TECNOLOGIA E NGENIERIA DE PRODUCCION B CAP(25)': 'Sala de Software B - 24 EST - 417 TECHNE', # Corregido error tipográfico 'NGENIERIA' si existe
            'LABORATORIO HAS CAP(22)': 'HAS-200 (317) TECHNE',
            'LABORATORIO FMS CAP(18)': 'FMS-200 (320) TECHNE',
            'LABORATORIO DE PROCESOS DE TRANSFORMACIÓN MECÁNICA': 'LABORATORIO DE PROCESOS DE TRANSFORMACIÓN BLOQUE 1-102'
            # Agrega más mapeos según sea necesario
            # 'NOMBRE_LABORATORIO_ORIGEN': 'NOMBRE_LABORATORIO_SALIDA',
        }
        
        # Días de la semana en orden
        self.dias = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO']
        
        # Franjas horarias en orden (6AM a 10PM)
        self.franjas_horarias = [
            '6AM-7AM', '7AM-8AM', '8AM-9AM', '9AM-10AM', '10AM-11AM',
            '11AM-12M', '12M-1PM', '1PM-2PM', '2PM-3PM', '3PM-4PM',
            '4PM-5PM', '5PM-6PM', '6PM-7PM', '7PM-8PM', '8PM-9PM', '9PM-10PM'
        ]
        
        # Columnas esperadas en el archivo de entrada
        self.columnas_entrada = [
            'Periodo', 'Día', 'Hora', 'Asignatura', 'Grupo', 
            'Proyecto', 'Salón', 'Área', 'Edificio', 'Sede', 
            'Inscritos', 'Docente'
        ]

    def actualizar_mapeo_laboratorios(self, nuevo_mapeo: Dict[str, str]):
        """
        Actualiza el diccionario de mapeo de nombres de laboratorios.
        
        Args:
            nuevo_mapeo: Diccionario con los mapeos de nombres de laboratorios.
        """
        self.mapeo_laboratorios.update(nuevo_mapeo)
        print("¡Mapeo de laboratorios actualizado exitosamente!")

    def leer_reporte_ocupacion(self, ruta_archivo: str) -> pd.DataFrame:
        """
        Lee el archivo de Excel del reporte de ocupación.
        
        Args:
            ruta_archivo: Ruta al archivo reporte_ocupacion.xlsx.
            
        Returns:
            DataFrame con los datos de ocupación.
        """
        try:
            if not os.path.exists(ruta_archivo):
                raise FileNotFoundError(f"Archivo no encontrado: {ruta_archivo}")
                
            df = pd.read_excel(ruta_archivo)
            
            # Validar columnas
            if len(df.columns) != len(self.columnas_entrada):
                print(f"Advertencia: Se esperaban {len(self.columnas_entrada)} columnas, pero se encontraron {len(df.columns)}")
                print(f"Esperadas: {self.columnas_entrada}")
                print(f"Encontradas: {list(df.columns)}")
            
            # Renombrar columnas para que coincidan con los nombres esperados
            df.columns = self.columnas_entrada[:len(df.columns)]
            
            print(f"Se cargaron exitosamente {len(df)} registros desde {ruta_archivo}")
            return df
            
        except Exception as e:
            raise Exception(f"Error al leer el reporte de ocupación: {str(e)}")

    def filtrar_laboratorios_mapeados(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Filtra el dataframe para incluir solo los laboratorios que están en nuestro mapeo y en el edificio TECHNE.
        
        Args:
            df: DataFrame de entrada.
            
        Returns:
            DataFrame filtrado solo con los laboratorios mapeados en el edificio TECHNE.
        """
        laboratorios_mapeados = list(self.mapeo_laboratorios.keys())
        
        # Filtrar por mapeo de laboratorios Y edificio = 'TECHNE'
        df_filtrado = df[
            (df['Salón'].isin(laboratorios_mapeados)) & 
            (df['Edificio'].str.upper() == 'TECHNE')
        ].copy()
        
        print(f"Se filtraron {len(df_filtrado)} registros para los laboratorios mapeados en el edificio TECHNE")
        print(f"Laboratorios mapeados encontrados: {df_filtrado['Salón'].unique().tolist()}")
        
        # Mostrar qué laboratorios fueron excluidos por el filtro de edificio
        todos_los_labs_mapeados = df[df['Salón'].isin(laboratorios_mapeados)]
        excluidos_por_edificio = todos_los_labs_mapeados[todos_los_labs_mapeados['Edificio'].str.upper() != 'TECHNE']
        if not excluidos_por_edificio.empty:
            labs_excluidos = excluidos_por_edificio['Salón'].unique().tolist()
            print(f"Laboratorios excluidos (no están en el edificio TECHNE): {labs_excluidos}")
        
        return df_filtrado

    def agrupar_horas_consecutivas(self, df: pd.DataFrame) -> List[Dict]:
        """
        Agrupa las entradas de horas consecutivas en sesiones de clase únicas.
        También maneja clases de una sola hora y clases no consecutivas.
        
        Args:
            df: DataFrame filtrado.
            
        Returns:
            Lista de diccionarios de sesiones de clase.
        """
        clases = []
        
        # Agrupar por todos los campos excepto la hora
        campos_agrupacion = ['Día', 'Asignatura', 'Grupo', 'Proyecto', 'Salón', 'Docente']
        agrupado = df.groupby(campos_agrupacion)
        
        for clave_grupo, df_grupo in agrupado:
            # Ordenar por hora para asegurar un orden consecutivo adecuado
            df_grupo = df_grupo.sort_values('Hora')
            horas = df_grupo['Hora'].tolist()
            horas_procesadas = set()
            
            # Encontrar primero los pares consecutivos
            i = 0
            while i < len(horas) - 1:
                if horas[i] in horas_procesadas:
                    i += 1
                    continue
                    
                hora_actual = horas[i]
                hora_siguiente = horas[i + 1]
                
                # Verificar si son horas consecutivas
                if self.son_horas_consecutivas(hora_actual, hora_siguiente):
                    info_clase = {
                        'dia': clave_grupo[0],
                        'hora_inicio': hora_actual,
                        'hora_fin': hora_siguiente,
                        'asignatura': clave_grupo[1],
                        'grupo': clave_grupo[2],
                        'proyecto': clave_grupo[3],
                        'laboratorio': clave_grupo[4],
                        'docente': clave_grupo[5],
                        'inscritos': df_grupo.iloc[0]['Inscritos'],  # Obtener el número de inscritos
                        'es_de_dos_horas': True
                    }
                    clases.append(info_clase)
                    horas_procesadas.add(hora_actual)
                    horas_procesadas.add(hora_siguiente)
                    i += 2  # Saltar la siguiente hora ya que ha sido procesada
                else:
                    i += 1
            
            # Manejar las horas restantes (clases de una hora o no consecutivas)
            for hora in horas:
                if hora not in horas_procesadas:
                    info_clase = {
                        'dia': clave_grupo[0],
                        'hora_inicio': hora,
                        'hora_fin': None,
                        'asignatura': clave_grupo[1],
                        'grupo': clave_grupo[2],
                        'proyecto': clave_grupo[3],
                        'laboratorio': clave_grupo[4],
                        'docente': clave_grupo[5],
                        'inscritos': df_grupo[df_grupo['Hora'] == hora].iloc[0]['Inscritos'],  # Obtener inscritos para esta hora
                        'es_de_dos_horas': False
                    }
                    clases.append(info_clase)
        
        print(f"Se encontraron {len(clases)} sesiones de clase (incluyendo clases de una hora)")
        return clases

    def son_horas_consecutivas(self, hora1: str, hora2: str) -> bool:
        """
        Verifica si dos horas son consecutivas en nuestra secuencia de franjas horarias.
        
        Args:
            hora1: Primera hora.
            hora2: Segunda hora.
            
        Returns:
            True si son consecutivas, False en caso contrario.
        """
        try:
            idx1 = self.franjas_horarias.index(hora1)
            idx2 = self.franjas_horarias.index(hora2)
            return idx2 == idx1 + 1
        except ValueError:
            return False

    def crear_matriz_horario(self, clases: List[Dict]) -> pd.DataFrame:
        """
        Crea la matriz de horario de salida con la estructura exacta de la plantilla.
        
        Args:
            clases: Lista de diccionarios de sesiones de clase.
            
        Returns:
            DataFrame con la matriz del horario.
        """
        # Obtener laboratorios únicos de nuestro mapeo (nombres de salida)
        laboratorios_salida = sorted(list(set(self.mapeo_laboratorios.values())))
        
        # Crear columnas: Dia, Hora, luego pares para cada laboratorio (Asignatura, Grupo)
        columnas = ['Dia', 'Hora']
        for lab in laboratorios_salida:
            columnas.extend([f"{lab}_asignatura", f"{lab}_grupo"])
        
        # Crear filas para cada día y franja horaria, más filas de separación
        filas = []
        for i, dia in enumerate(self.dias):
            for franja in self.franjas_horarias:
                fila = {'Dia': dia, 'Hora': franja}
                # Inicializar todas las columnas de laboratorio como vacías
                for lab in laboratorios_salida:
                    fila[f"{lab}_asignatura"] = ''
                    fila[f"{lab}_grupo"] = ''
                filas.append(fila)
            
            # Añadir fila de separación después de cada día (excepto el último)
            if i < len(self.dias) - 1:
                fila_separadora = {'Dia': '', 'Hora': ''}
                for lab in laboratorios_salida:
                    fila_separadora[f"{lab}_asignatura"] = ''
                    fila_separadora[f"{lab}_grupo"] = ''
                filas.append(fila_separadora)
        
        # Crear el dataframe base
        df_horario = pd.DataFrame(filas, columns=columnas)
        
        # Rellenar la información de las clases
        for info_clase in clases:
            dia = info_clase['dia']
            hora_inicio = info_clase['hora_inicio']
            
            # Mapear nombre de laboratorio a nombre de salida
            laboratorio_salida = self.mapeo_laboratorios.get(info_clase['laboratorio'])
            if not laboratorio_salida:
                continue
            
            # Crear grupo con información de inscritos
            grupo_con_inscritos = f"{info_clase['grupo']} | {info_clase['inscritos']}"
            
            if info_clase['es_de_dos_horas']:
                # Clase de dos horas: formato tradicional
                hora_fin = info_clase['hora_fin']
                
                # Primera hora: Asignatura en la primera columna, Grupo con inscritos en la segunda
                mascara_inicio = (df_horario['Dia'] == dia) & (df_horario['Hora'] == hora_inicio)
                if mascara_inicio.any():
                    idx = df_horario[mascara_inicio].index[0]
                    df_horario.loc[idx, f"{laboratorio_salida}_asignatura"] = info_clase['asignatura']
                    df_horario.loc[idx, f"{laboratorio_salida}_grupo"] = grupo_con_inscritos
                
                # Segunda hora: Docente en la primera columna, Proyecto en la segunda
                mascara_fin = (df_horario['Dia'] == dia) & (df_horario['Hora'] == hora_fin)
                if mascara_fin.any():
                    idx = df_horario[mascara_fin].index[0]
                    df_horario.loc[idx, f"{laboratorio_salida}_asignatura"] = info_clase['docente']
                    df_horario.loc[idx, f"{laboratorio_salida}_grupo"] = info_clase['proyecto']
            else:
                # Clase de una hora: poner asignatura y grupo en la primera fila
                mascara_inicio = (df_horario['Dia'] == dia) & (df_horario['Hora'] == hora_inicio)
                if mascara_inicio.any():
                    idx = df_horario[mascara_inicio].index[0]
                    df_horario.loc[idx, f"{laboratorio_salida}_asignatura"] = info_clase['asignatura']
                    df_horario.loc[idx, f"{laboratorio_salida}_grupo"] = grupo_con_inscritos
        
        return df_horario

    def formatear_encabezados_salida(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
        """
        Formatea los encabezados de salida para que coincidan con la estructura de la plantilla.
        
        Args:
            df: DataFrame del horario.
            
        Returns:
            Un tuple con el DataFrame con encabezados formateados y la lista de nombres de laboratorios.
        """
        # Crear una copia para evitar modificar el original
        df_salida = df.copy()
        
        # Crear nuevos nombres de columna para mayor legibilidad
        nuevas_columnas = []
        nombres_laboratorios = []
        
        for col in df_salida.columns:
            if col in ['Dia', 'Hora']:
                nuevas_columnas.append(col)
            elif col.endswith('_asignatura'):
                nombre_lab = col.replace('_asignatura', '')
                nuevas_columnas.append(f"{nombre_lab} - Asignatura")
                if nombre_lab not in nombres_laboratorios:
                    nombres_laboratorios.append(nombre_lab)
            elif col.endswith('_grupo'):
                nombre_lab = col.replace('_grupo', '')
                nuevas_columnas.append(f"{nombre_lab} - Grupo")
        
        df_salida.columns = nuevas_columnas
        
        return df_salida, nombres_laboratorios

    def guardar_horario(self, df: pd.DataFrame, ruta_salida: str):
        """
        Guarda el horario en un archivo de Excel con formato adecuado, separaciones por día y colores alternos.
        
        Args:
            df: DataFrame del horario.
            ruta_salida: Ruta del archivo de salida.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Crear libro y hoja de trabajo
            libro = Workbook()
            hoja = libro.active
            hoja.title = "Horario Laboratorios"
            
            # Escribir el dataframe en la hoja de trabajo
            for r in dataframe_to_rows(df, index=False, header=True):
                hoja.append(r)
            
            # Definir colores y estilos
            relleno_encabezado = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            fuente_encabezado = Font(color="FFFFFF", bold=True)
            relleno_separador = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            # Colores alternos para pares de filas (verde y azul)
            relleno_verde = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # Verde claro
            relleno_azul = PatternFill(start_color="E8F0FF", end_color="E8F0FF", fill_type="solid")   # Azul claro
            
            # Formatear encabezados
            for col in range(1, hoja.max_column + 1):
                celda = hoja.cell(row=1, column=col)
                celda.fill = relleno_encabezado
                celda.font = fuente_encabezado
                celda.alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar colores alternos a pares de filas y manejar separadores
            for fila_num in range(2, hoja.max_row + 1):
                # Comprobar si es una fila separadora (Dia y Hora vacíos)
                celda_dia = hoja.cell(row=fila_num, column=1)
                celda_hora = hoja.cell(row=fila_num, column=2)
                
                if not celda_dia.value and not celda_hora.value:
                    # Es una fila separadora
                    for col in range(1, hoja.max_column + 1):
                        hoja.cell(row=fila_num, column=col).fill = relleno_separador
                else:
                    # Es una fila de datos normal
                    franja_horaria = celda_hora.value
                    
                    if franja_horaria in self.franjas_horarias:
                        indice_tiempo = self.franjas_horarias.index(franja_horaria)
                        indice_par = indice_tiempo // 2  # División entera para obtener el número de par
                        
                        # Alternar colores para cada par
                        color_relleno = relleno_verde if indice_par % 2 == 0 else relleno_azul
                        
                        for col in range(1, hoja.max_column + 1):
                            hoja.cell(row=fila_num, column=col).fill = color_relleno
            
            # Añadir bordes a todas las celdas
            borde_fino = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for fila in range(1, hoja.max_row + 1):
                for col in range(1, hoja.max_column + 1):
                    celda = hoja.cell(row=fila, column=col)
                    celda.border = borde_fino
                    celda.alignment = Alignment(
                        horizontal="center", 
                        vertical="center",
                        wrap_text=True
                    )
            
            # Autoajustar anchos de columna
            for columna in hoja.columns:
                longitud_maxima = 0
                letra_columna = columna[0].column_letter
                for celda in columna:
                    try:
                        if len(str(celda.value)) > longitud_maxima:
                            longitud_maxima = len(str(celda.value))
                    except:
                        pass
                ancho_ajustado = min(longitud_maxima + 2, 50)  # Limitar a 50 caracteres
                hoja.column_dimensions[letra_columna].width = ancho_ajustado
            
            # Guardar el libro de trabajo
            libro.save(ruta_salida)
            
            print(f"Horario guardado exitosamente en: {ruta_salida}")
            
        except Exception as e:
            # Opción de respaldo: guardado básico con pandas si falla el formato con openpyxl
            print(f"Falló el formato avanzado, guardando versión básica: {str(e)}")
            df.to_excel(ruta_salida, index=False, sheet_name='Horario')
            print(f"Horario básico guardado en: {ruta_salida}")

    def generar_horario(self, archivo_entrada: str, archivo_salida: str):
        """
        Método principal para generar el horario completo.
        
        Args:
            archivo_entrada: Ruta a reporte_ocupacion.xlsx.
            archivo_salida: Ruta para el archivo de salida HORARIO_LABORATORIOS.xlsx.
        """
        print("Iniciando la generación de horarios de laboratorio...")
        print(f"Archivo de entrada: {archivo_entrada}")
        print(f"Archivo de salida: {archivo_salida}")
        print(f"Mapeos de laboratorio: {self.mapeo_laboratorios}")
        
        try:
            # Paso 1: Leer el reporte de ocupación
            df = self.leer_reporte_ocupacion(archivo_entrada)
            
            # Paso 2: Filtrar por laboratorios mapeados
            df_filtrado = self.filtrar_laboratorios_mapeados(df)
            
            if df_filtrado.empty:
                print("Advertencia: ¡No se encontraron datos para los laboratorios mapeados!")
                return
            
            # Paso 3: Agrupar horas consecutivas en clases
            clases = self.agrupar_horas_consecutivas(df_filtrado)
            
            if not clases:
                print("Advertencia: ¡No se encontraron sesiones de clase completas!")
                return
            
            # Paso 4: Crear la matriz del horario
            df_horario = self.crear_matriz_horario(clases)
            
            # Paso 5: Formatear encabezados
            df_formateado, _ = self.formatear_encabezados_salida(df_horario)
            
            # Paso 6: Guardar la salida
            self.guardar_horario(df_formateado, archivo_salida)
            
            print("¡La generación de horarios de laboratorio se completó exitosamente!")
            
        except Exception as e:
            print(f"Error durante la generación del horario: {str(e)}")
            raise

def principal():
    """
    Ejemplo de uso del Generador de Horarios de Laboratorio.
    """
    # Crear la instancia del generador
    generador = GeneradorHorariosLaboratorio()
    
    # Opcional: Actualizar mapeos de laboratorios si es necesario
    mapeos_adicionales = {
        # 'NUEVO_LABORATORIO_ORIGEN': 'NUEVO_LABORATORIO_SALIDA',
    }
    if mapeos_adicionales:
        generador.actualizar_mapeo_laboratorios(mapeos_adicionales)
    
    # Rutas de los archivos
    archivo_entrada = "reporte_ocupacion.xlsx"
    archivo_salida = "HORARIO_LABORATORIOS.xlsx"
    
    # Generar el horario
    try:
        generador.generar_horario(archivo_entrada, archivo_salida)
    except Exception as e:
        print(f"Falló la generación del horario: {str(e)}")

if __name__ == "__main__":
    principal()